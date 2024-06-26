from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter

# make the worksheet for totality
wb = Workbook()
sheet_total = wb.active
sheet_total.title = 'total'

# insert column titles
column_titles = ['ASIN', '品名', '价格', '曝光', '点击', 'CTR',
                 'ACOS',
                 'ACOAS', '广告费',
                 '店铺销售额', '广告销售额', '店铺销量', '广告销量',
                 '毛利润', '毛利率', '排名']


# copy titles to the sheet_total
def copy_titles(titles_list, sheet):
    column_titles_len = len(titles_list)
    count = 0
    for row in sheet.iter_rows(min_row=1,
                               max_col=column_titles_len):
        for cell in row:
            if count < column_titles_len:
                cell.value = column_titles[count]
            count += 1


# fill sheet totals
copy_titles(column_titles, sheet_total)

# clean up source 1 sheet
source1 = load_workbook('sample1.xlsx')
source_sheet1 = source1.active
source_sheet1.delete_rows(2, amount=1)


# make a function to copy from source1 to sheet_total
def copy_source1(source_row_num, dest_col):
    list_source = []
    for i, row in enumerate(source_sheet1):
        if i == 0:
            continue
        item = row[source_row_num].value
        list_source.append(item)
    list_len = len(list_source)
    i = 0
    for row in sheet_total[f"{dest_col}2:{dest_col}{list_len + 1}"]:
        for cell in row:
            cell.value = list_source[i]
            i += 1


# copy asin col=3
copy_source1(3, "A")

# copy 价格/竞价 col=7
copy_source1(7, "C")

# copy 曝光 col=9
copy_source1(9, "D")

# copy 点击 col=10
copy_source1(10, "E")

# copy CTR col=11
copy_source1(11, "F")

# copy ACOS col=17
copy_source1(17, "G")

# copy ACOAS col=18
copy_source1(18, "H")

# copy 广告花费 col=13
copy_source1(13, "I")

# copy 店铺销售额 col=14
copy_source1(14, "J")

# copy 广告销售额 col=14
copy_source1(15, "K")

# copy 店铺销量 col=23
copy_source1(23, "L")

# copy 广告销量 col=24
copy_source1(24, "M")

# close source sheet1
# source1.save("sample1.xlsx")

# clean up source 2 sheet
source2 = load_workbook('sample2.xlsx')
source_sheet2 = source2.active
source_sheet2.delete_rows(1, amount=1)


# #make a function to find index of asin
# def find_index(source_sheet,lookup_value, lookup_col_num):
#     source_list=[]
#     for i,row in enumerate(source_sheet):
#         if i==0:
#             continue
#         item=row[lookup_col_num].value
#         source_list.append(item)
#     target_index=source_list.index(lookup_value)
#     return target_index
#
# #make a function to return the value of the look up value
# def return_value(target_index,return_row_num):
#     return_list=[]
#     for i,row in enumerate(source_sheet2):
#         if i==0:
#             continue
#         item=row[return_row_num].value
#         return_list.append(item)
#         return(return_list[target_index])

# make a function the generate a dictionary of key_asin and
# value_target value.
def make_dict_source(source_sheet, lookup_col_num, return_col_num):
    dict = {}
    for i, row in enumerate(source_sheet):
        if i == 0:
            continue
        key = row[lookup_col_num].value
        value = row[return_col_num].value
        dict[key] = value

    return dict


# print(make_dict_source2(1,3))

# fill the sheet_total with the dict from source2
def fill_sheet_total(source_dict,
                     lookup_col_num,
                     target_col_num):
    source_dict = source_dict
    for row in sheet_total:
        if row[target_col_num].value == None:
            row[target_col_num].value = source_dict.get(
                row[lookup_col_num].value)


# fill 品名
fill_sheet_total(make_dict_source(source_sheet2, 1, 2), 0, 1)

# fill 毛利润
fill_sheet_total(make_dict_source(source_sheet2, 1, 3), 0, 13)

# fill 毛利率
fill_sheet_total(make_dict_source(source_sheet2, 1, 4), 0, 14)

# close source sheet 2
# source2.save("sample2.xlsx")


# load source sheet 3
source3 = load_workbook('sample3.xlsx')
source_sheet3 = source3.active


def edit_source_dict(source_dict):
    ready_dict = source_dict
    modified_dict = {}
    for key, value in ready_dict.items():
        if value == None:
            modified_dict[key] = 0
        else:
            modified_dict[key] = value.split('：')[-1]

    return modified_dict


# fill 排名
fill_sheet_total(edit_source_dict(make_dict_source(source_sheet3, 1,
                                                   2)), 0, 15)

# total sheet for myyweld
sheet_mw = wb.create_sheet("MW")
# total sheet for ttamplar
sheet_tp = wb.create_sheet("TP")

# copy titles to two sheets
copy_titles(column_titles, sheet_mw)
copy_titles(column_titles, sheet_tp)


# sort the data to the two sheets function
def sort_data():
    for i, row in enumerate(sheet_total):
        if i == 0:
            continue
        if 'W' in row[1].value:
            mw_max_row = sheet_mw.max_row + 1
            for cell in row:
                sheet_mw.cell(row=mw_max_row,
                              column=cell.column).value = cell.value
        else:
            tp_max_row = sheet_tp.max_row + 1
            for cell in row:
                sheet_tp.cell(row=tp_max_row,
                              column=cell.column).value = cell.value

#sort and fill the two brand sheets.
sort_data()

#sum up the two brand sheet.
def sum_up(sheet):
    max_row = sheet.max_row
    for cell in sheet[max_row+1]:
        if 3<cell.column<15:
            column_letter=get_column_letter(cell.column)
            print(column_letter)
            cell.value = (f"=SUM({column_letter}2:{column_letter}"
                          f"{max_row})")

#sum up for two brand sheets
sum_up(sheet_mw)
sum_up(sheet_tp)

# save the final sheet_total
wb.save('total.xlsx')
